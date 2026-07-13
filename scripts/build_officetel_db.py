#!/usr/bin/env python3
"""
build_officetel_db.py — 국세청 오피스텔 기준시가 xlsx → 조회용 SQLite 적재

연립·다세대는 VWorld 실시간 API라 적재가 필요 없지만,
오피스텔은 국세청 파일이라 한 번 적재해 두면 PNU+호로 빠르게 조회된다.

[입력] 국세청 상업용건물/오피스텔 기준시가 xlsx (연도별)
   - 시트 1~5에 나뉘어 있음 (총 ~249만 행, 그중 오피스텔 ~133만)
   - 상가종류코드='오피스텔'만 적재
[출력] SQLite officetel 테이블 (pnu, ho, year 등) + (pnu,ho) 인덱스

바이브코더용 사용법 (PowerShell/터미널):
   python build_officetel_db.py --xlsx 2026파일.xlsx --year 2026 --db officetel.db
   python build_officetel_db.py --xlsx 2025파일.xlsx --year 2025 --db officetel.db
   (같은 db에 연도별로 두 번 돌리면 구·신 비교 가능)

전국 전부 적재하려면 --region 생략. 특정 시도만 하려면 --region 11 41 …
"""

import argparse
import os
import sqlite3
import sys

# build_pnu를 패키지에서 가져온다(같은 PNU 규칙 공유)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from jeonse_pnu import build_pnu  # noqa: E402


def ensure_schema(con):
    # price=고시가격(㎡당 단가). 총액은 조회 시 price×(prvuse+share)로 계산.
    # PK: 같은 PNU에 여러 동/층/호가 있어 (pnu,ho,year)만으로는 덮어씀 → 확장.
    con.execute("""CREATE TABLE IF NOT EXISTS officetel(
        pnu TEXT, ldcode TEXT, building TEXT, dong TEXT, floor TEXT, ho TEXT,
        prvuse REAL, share REAL, price INTEGER, year TEXT,
        PRIMARY KEY(pnu, building, dong, floor, ho, year))""")
    con.execute("CREATE INDEX IF NOT EXISTS idx_off_pnu_ho ON officetel(pnu, ho)")
    con.commit()


def load(xlsx_path, year, db_path, regions=None, batch=5000):
    import openpyxl
    if not os.path.exists(xlsx_path):
        sys.exit(f"[오류] 파일 없음: {xlsx_path}")

    con = sqlite3.connect(db_path)
    ensure_schema(con)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    n, skipped, buf = 0, 0, []
    region_prefixes = tuple(regions) if regions else None

    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            # 컬럼: 0상가건물번호 1상가종류코드 2고시일자 3법정동코드 4특수지코드
            #       5번지 6호 7블록주소 8동주소 9층구분 10층주소 11호주소
            #       12고시가격 13전용면적 14공유면적
            if row[1] != "오피스텔":
                continue
            ldcode = str(row[3])
            if region_prefixes and not ldcode.startswith(region_prefixes):
                continue
            san = "1" if "산" in str(row[4]) else "0"
            try:
                pnu = build_pnu(ldcode, bonbun=row[5], bubun=row[6],
                                mountain=san, mountain_source="juso")
            except Exception:
                skipped += 1
                continue
            # 컬럼: 12고시가격(㎡당) 13전용면적 14공유면적
            buf.append((pnu, ldcode, row[7], str(row[8]), str(row[10]),
                        str(row[11]), row[13], row[14], row[12], str(year)))
            if len(buf) >= batch:
                con.executemany(
                    "INSERT OR REPLACE INTO officetel VALUES (?,?,?,?,?,?,?,?,?,?)", buf)
                con.commit()
                n += len(buf)
                buf.clear()
                print(f"\r  적재 중... {n:,}행", end="", flush=True)
    if buf:
        con.executemany(
            "INSERT OR REPLACE INTO officetel VALUES (?,?,?,?,?,?,?,?,?,?)", buf)
        con.commit()
        n += len(buf)
    wb.close()

    total = con.execute("SELECT COUNT(*) FROM officetel WHERE year=?", (str(year),)).fetchone()[0]
    pnus = con.execute("SELECT COUNT(DISTINCT pnu) FROM officetel WHERE year=?", (str(year),)).fetchone()[0]
    con.close()
    print(f"\n[완료] {year}년 {n:,}행 적재 (스킵 {skipped}) "
          f"| DB내 {year}년 총 {total:,}행 / 고유 PNU {pnus:,}개")


def main():
    ap = argparse.ArgumentParser(description="오피스텔 기준시가 xlsx → SQLite 적재")
    ap.add_argument("--xlsx", required=True, help="국세청 오피스텔 기준시가 xlsx 경로")
    ap.add_argument("--year", required=True, help="기준연도 (예: 2026)")
    ap.add_argument("--db", default="officetel.db", help="출력 SQLite 경로")
    ap.add_argument("--region", nargs="*", default=None,
                    help="적재할 법정동코드 접두(생략=전국). 예: --region 11 41")
    args = ap.parse_args()
    load(args.xlsx, args.year, args.db, regions=args.region)


if __name__ == "__main__":
    main()
