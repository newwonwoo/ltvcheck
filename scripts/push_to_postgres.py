#!/usr/bin/env python3
"""
push_to_postgres.py — 오피스텔 기준시가를 Vercel Postgres(Neon)에 적재

두 가지 방식 지원:
  (A) 이미 만든 SQLite를 Postgres로 옮기기 (권장 — 빠름)
      python push_to_postgres.py --from-sqlite officetel.db
  (B) xlsx에서 바로 Postgres로 적재
      python push_to_postgres.py --xlsx 2026파일.xlsx --year 2026
      python push_to_postgres.py --xlsx 2025파일.xlsx --year 2025

연결 정보(POSTGRES_URL)는 환경변수로 준다. Vercel에서 Postgres 생성 후
대시보드 > Storage > .env.local 다운로드하면 POSTGRES_URL이 들어있다.
   예) export POSTGRES_URL="postgres://user:pass@ep-xxx.neon.tech/db?sslmode=require"

바이브코더용 요약:
  1) Vercel > Storage > Create Database > Postgres(Neon) 생성
  2) 그 화면에서 POSTGRES_URL 복사 → 로컬 환경변수로 설정
  3) 이 스크립트로 적재 (A 또는 B)
  4) 끝 — Vercel 함수는 POSTGRES_URL을 자동으로 읽어 오피스텔 조회 가능
"""

import argparse
import os
import sqlite3
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from jeonse_pnu import build_pnu  # noqa: E402

DDL = """
CREATE TABLE IF NOT EXISTS officetel(
    pnu TEXT, ldcode TEXT, building TEXT, dong TEXT, floor TEXT, ho TEXT,
    prvuse REAL, price BIGINT, year TEXT,
    PRIMARY KEY(pnu, ho, year)
);
CREATE INDEX IF NOT EXISTS idx_off_pnu_ho ON officetel(pnu, ho);
"""

UPSERT = """
INSERT INTO officetel(pnu, ldcode, building, dong, floor, ho, prvuse, price, year)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
ON CONFLICT (pnu, ho, year) DO UPDATE SET
    building=EXCLUDED.building, dong=EXCLUDED.dong, floor=EXCLUDED.floor,
    prvuse=EXCLUDED.prvuse, price=EXCLUDED.price
"""


def _connect():
    url = os.environ.get("POSTGRES_URL", "").strip() or os.environ.get("DATABASE_URL", "").strip()
    if not url:
        sys.exit("[오류] POSTGRES_URL 환경변수가 없습니다. Vercel Storage에서 복사해 설정하세요.")
    try:
        import psycopg
    except ImportError:
        sys.exit("[오류] psycopg 미설치. pip install 'psycopg[binary]'")
    return psycopg.connect(url)


def ensure_schema(conn):
    with conn.cursor() as cur:
        cur.execute(DDL)
    conn.commit()


def _flush(conn, buf):
    with conn.cursor() as cur:
        cur.executemany(UPSERT, buf)
    conn.commit()


def from_sqlite(sqlite_path, batch=5000):
    """이미 만든 SQLite officetel 테이블 → Postgres 복사."""
    if not os.path.exists(sqlite_path):
        sys.exit(f"[오류] SQLite 없음: {sqlite_path}")
    conn = _connect()
    ensure_schema(conn)
    src = sqlite3.connect(sqlite_path)
    src.row_factory = sqlite3.Row
    cols = "pnu, ldcode, building, dong, floor, ho, prvuse, price, year"
    rows = src.execute(f"SELECT {cols} FROM officetel")
    n, buf = 0, []
    for r in rows:
        buf.append((r["pnu"], r["ldcode"], r["building"], r["dong"], r["floor"],
                    r["ho"], r["prvuse"], r["price"], r["year"]))
        if len(buf) >= batch:
            _flush(conn, buf); n += len(buf); buf.clear()
            print(f"\r  적재 중... {n:,}행", end="", flush=True)
    if buf:
        _flush(conn, buf); n += len(buf)
    src.close()
    _report(conn, n)


def from_xlsx(xlsx_path, year, regions=None, batch=5000):
    """국세청 xlsx → Postgres 직접 적재 (build_officetel_db.py와 동일 파싱)."""
    import openpyxl
    if not os.path.exists(xlsx_path):
        sys.exit(f"[오류] 파일 없음: {xlsx_path}")
    conn = _connect()
    ensure_schema(conn)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    prefixes = tuple(regions) if regions else None
    n, skipped, buf = 0, 0, []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row[1] != "오피스텔":
                continue
            ldcode = str(row[3])
            if prefixes and not ldcode.startswith(prefixes):
                continue
            san = "1" if "산" in str(row[4]) else "0"
            try:
                pnu = build_pnu(ldcode, bonbun=row[5], bubun=row[6],
                                mountain=san, mountain_source="juso")
            except Exception:
                skipped += 1
                continue
            buf.append((pnu, ldcode, row[7], str(row[8]), str(row[10]),
                        str(row[11]), row[13], row[12], str(year)))
            if len(buf) >= batch:
                _flush(conn, buf); n += len(buf); buf.clear()
                print(f"\r  적재 중... {n:,}행", end="", flush=True)
    if buf:
        _flush(conn, buf); n += len(buf)
    wb.close()
    print(f"\n  (스킵 {skipped}행)")
    _report(conn, n)


def _report(conn, n):
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*), COUNT(DISTINCT pnu) FROM officetel")
        total, pnus = cur.fetchone()
    conn.close()
    print(f"\n[완료] {n:,}행 적재 | Postgres 총 {total:,}행 / 고유 PNU {pnus:,}개")


def main():
    ap = argparse.ArgumentParser(description="오피스텔 기준시가 → Vercel Postgres(Neon) 적재")
    ap.add_argument("--from-sqlite", help="이미 만든 SQLite 경로(권장)")
    ap.add_argument("--xlsx", help="국세청 오피스텔 기준시가 xlsx")
    ap.add_argument("--year", help="기준연도 (xlsx 방식일 때 필수)")
    ap.add_argument("--region", nargs="*", default=None,
                    help="법정동코드 접두 필터(생략=전국)")
    args = ap.parse_args()

    if args.from_sqlite:
        from_sqlite(args.from_sqlite)
    elif args.xlsx:
        if not args.year:
            sys.exit("[오류] --xlsx 방식은 --year 필요")
        from_xlsx(args.xlsx, args.year, regions=args.region)
    else:
        sys.exit("[사용법] --from-sqlite officetel.db  또는  --xlsx 파일 --year 2026")


if __name__ == "__main__":
    main()
